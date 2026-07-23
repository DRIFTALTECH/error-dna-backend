"""Extract text from downloaded note attachments, to feed the LLM alongside the note body.

Text-native files (xsd/xml/txt/json/log/csv/...) are read directly; PDFs via pypdf;
docx via python-docx; xlsx via openpyxl. Anything else (images, binaries) → "" (skipped).
All best-effort: a failed extract returns "" and never raises.
"""

import os
import logging

logger = logging.getLogger(__name__)

# Read these as plain text.
TEXT_EXTS = {
    ".xsd", ".xml", ".wsdl", ".txt", ".json", ".log", ".csv", ".tsv",
    ".properties", ".groovy", ".java", ".sql", ".yaml", ".yml", ".md", ".html",
}
SUPPORTED_EXTS = TEXT_EXTS | {".pdf", ".docx", ".xlsx"}
MAX_ATTACH_CHARS = 15000  # cap total attachment text fed to the LLM


def extract_text(path: str) -> str:
    """Return extracted text for a file, or '' if unsupported/unreadable."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in TEXT_EXTS:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(path)
            return "\n".join((pg.extract_text() or "") for pg in reader.pages)
        if ext == ".docx":
            import docx
            doc = docx.Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows.append("\t".join(cells))
            return "\n".join(rows)
    except Exception as e:
        logger.warning(f"attachment extract failed for {os.path.basename(path)}: {e}")
        return ""
    return ""  # images / unknown → skip


if __name__ == "__main__":
    import sys
    if len(sys.argv) == 2:
        t = extract_text(sys.argv[1])
        print(f"[{len(t)} chars]\n{t[:800]}")
    else:
        print("usage: python3 -m services.attachments <file>")
